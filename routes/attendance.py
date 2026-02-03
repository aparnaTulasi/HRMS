from flask import Blueprint, request, jsonify, g
from datetime import datetime, date
import csv
import io

from models import db
from models.attendance import Attendance, AttendanceRegularization
from models.employee import Employee
from models.user import User
from utils.decorators import token_required, role_required

attendance_bp = Blueprint("attendance", __name__)

ALLOWED_MANAGE_ROLES = ["SUPER_ADMIN", "ADMIN", "HR"]


# -----------------------------
# Helpers
# -----------------------------
def _parse_date(value: str) -> date:
    """
    Accepts:
      - YYYY-MM-DD
      - DD/MM/YYYY
      - MM/DD/YYYY (if you want; can remove)
    """
    if not value:
        raise ValueError("date is required")

    value = value.strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(value, fmt).date()
        except ValueError:
            continue
    raise ValueError(f"Invalid date format: {value}")


def _parse_time(value: str, base_date: date) -> datetime:
    """
    Accepts:
      - HH:MM
      - HH:MM:SS
      - 09:45
      - 19:34
    """
    if not value:
        return None

    value = value.strip()
    for fmt in ("%H:%M", "%H:%M:%S"):
        try:
            t = datetime.strptime(value, fmt).time()
            return datetime.combine(base_date, t)
        except ValueError:
            continue
    raise ValueError(f"Invalid time format: {value}")


def _format_logged_time(total_minutes: int) -> str:
    hrs = total_minutes // 60
    mins = total_minutes % 60
    if hrs > 0 and mins > 0:
        return f"{hrs} hrs {mins} mins"
    if hrs > 0:
        return f"{hrs} hrs"
    return f"{mins} mins"


def _get_employee_in_company(employee_id: int) -> Employee:
    emp = Employee.query.get(employee_id)
    if not emp or emp.company_id != g.user.company_id:
        return None
    return emp


def _upsert_attendance(company_id: int, employee_id: int, att_date: date, payload: dict, capture_method: str):
    row = Attendance.query.filter_by(
        company_id=company_id,
        employee_id=employee_id,
        attendance_date=att_date
    ).first()

    is_new = False
    if not row:
        row = Attendance(company_id=company_id, employee_id=employee_id, attendance_date=att_date)
        row.created_by = getattr(g.user, "id", None)
        db.session.add(row)
        is_new = True

    # 1. Update status if provided
    if payload.get("status"):
        row.status = payload["status"]

    # 2. If Absent, CLEAR times. Otherwise, update times if provided.
    if row.status == "Absent":
        row.punch_in_time = None
        row.punch_out_time = None
    else:
        if payload.get("login_at") is not None:
            row.punch_in_time = payload["login_at"]
        if payload.get("logout_at") is not None:
            row.punch_out_time = payload["logout_at"]

    row.capture_method = capture_method

    row.updated_by = getattr(g.user, "id", None)

    # Auto calculate logged time
    row.recalc_total_minutes()

    return row, is_new


# -----------------------------
# 1) List Attendance (Table)
# -----------------------------
@attendance_bp.route("", methods=["GET"])
@token_required
@role_required(ALLOWED_MANAGE_ROLES)
def list_attendance():
    """
    Filters used by your UI:
      - role (Admin/HR/Manager/Employee/Accountant)
      - department
      - day=all/today
      - month (1-12) optional
      - from_date, to_date
      - search (name/email)
    """
    company_id = g.user.company_id

    role = request.args.get("role")          # Role dropdown
    department = request.args.get("department")
    day = (request.args.get("day") or "all").lower()  # all / today
    month = request.args.get("month")        # 1..12
    from_date = request.args.get("from_date")
    to_date = request.args.get("to_date")
    search = (request.args.get("search") or "").strip().lower()

    q = Attendance.query.filter_by(company_id=company_id)

    # Day filter
    if day == "today":
        q = q.filter(Attendance.attendance_date == date.today())

    # Date range filter
    if from_date:
        q = q.filter(Attendance.attendance_date >= _parse_date(from_date))
    if to_date:
        q = q.filter(Attendance.attendance_date <= _parse_date(to_date))

    # Month filter (optional)
    if month:
        try:
            m = int(month)
            q = q.filter(db.extract("month", Attendance.attendance_date) == m)
        except:
            pass

    # Join with employee + user for role/department/search
    q = q.join(Employee, Employee.id == Attendance.employee_id)         .join(User, User.id == Employee.user_id)         .filter(Employee.company_id == company_id)

    if department:
        q = q.filter(Employee.department == department)

    if role:
        q = q.filter(User.role == role)

    if search:
        q = q.filter(
            db.or_(
                db.func.lower(Employee.first_name).like(f"%{search}%"),
                db.func.lower(Employee.last_name).like(f"%{search}%"),
                db.func.lower(User.email).like(f"%{search}%"),
                db.func.lower(getattr(Employee, "employee_id", "")).like(f"%{search}%")
            )
        )

    rows = q.order_by(Attendance.attendance_date.desc()).limit(500).all()

    output = []
    for r in rows:
        emp = Employee.query.get(r.employee_id)
        user = User.query.get(emp.user_id) if emp else None

        output.append({
            "attendance_id": r.attendance_id,
            "employee_id": emp.employee_id if emp else "",
            "name": f"{emp.first_name} {emp.last_name}" if emp else "",
            "role": user.role if user else None,
            "department": emp.department if emp else None,
            "status": r.status,
            "logged_time": _format_logged_time(r.total_minutes),
            "login_at": r.punch_in_time.strftime("%H:%M") if r.punch_in_time else None,
            "logout_at": r.punch_out_time.strftime("%H:%M") if r.punch_out_time else None,
            "date": r.attendance_date.strftime("%d/%m/%Y"),
        })

    return jsonify({"attendance": output}), 200


# -----------------------------
# 2) Manual Attendance (Create/Upsert)
# -----------------------------
@attendance_bp.route("/manual", methods=["POST"])
@token_required
@role_required(ALLOWED_MANAGE_ROLES)
def manual_attendance():
    """
    Manual button action (UPSERT):
      Required: employee_id, date
      Optional: status, login_at, logout_at
    """
    data = request.get_json() or {}

    employee_id = data.get("employee_id")
    att_date = data.get("date")  # "2025-10-02" or "02/10/2025"
    status = data.get("status")  # Don't default to "Present" here, let upsert handle it

    if not employee_id or not att_date:
        return jsonify({"message": "employee_id and date are required"}), 400

    # Handle Super Admin lookup (find employee across any company)
    query = Employee.query.filter_by(employee_id=employee_id)
    if g.user.role != 'SUPER_ADMIN':
        query = query.filter_by(company_id=g.user.company_id)
    emp = query.first()
    if not emp:
        return jsonify({"message": "Employee not found"}), 404

    d = _parse_date(att_date)

    login_at = None
    logout_at = None

    # If absent -> times should be empty (allowed)
    if data.get("login_at"):
        login_at = _parse_time(data["login_at"], d)
    if data.get("logout_at"):
        logout_at = _parse_time(data["logout_at"], d)

    payload = {
        "status": status,
        "login_at": login_at,
        "logout_at": logout_at
    }

    row, is_new = _upsert_attendance(emp.company_id, emp.id, d, payload, capture_method="Manual")
    db.session.commit()

    return jsonify({
        "message": "Attendance saved successfully",
        "action": "inserted" if is_new else "updated",
        "attendance_id": row.attendance_id
    }), 200


# -----------------------------
# 3) Update Attendance (Edit icon)
# -----------------------------
@attendance_bp.route("/<int:attendance_id>", methods=["PUT"])
@token_required
@role_required(ALLOWED_MANAGE_ROLES)
def update_attendance(attendance_id):
    data = request.get_json() or {}
    row = Attendance.query.get(attendance_id)

    if not row or row.company_id != g.user.company_id:
        return jsonify({"message": "Attendance not found"}), 404

    # Update values
    if "status" in data:
        row.status = data["status"] or row.status

    if "login_at" in data:
        if data["login_at"]:
            row.punch_in_time = _parse_time(data["login_at"], row.attendance_date)
        else:
            row.punch_in_time = None

    if "logout_at" in data:
        if data["logout_at"]:
            row.punch_out_time = _parse_time(data["logout_at"], row.attendance_date)
        else:
            row.punch_out_time = None

    row.capture_method = "Manual"
    row.updated_by = getattr(g.user, "id", None)
    row.recalc_total_minutes()

    db.session.commit()
    return jsonify({"message": "Attendance updated successfully"}), 200


# -----------------------------
# 4) Delete Attendance (Delete icon)
# -----------------------------
@attendance_bp.route("/<int:attendance_id>", methods=["DELETE"])
@token_required
@role_required(ALLOWED_MANAGE_ROLES)
def delete_attendance(attendance_id):
    row = Attendance.query.get(attendance_id)
    if not row or row.company_id != g.user.company_id:
        return jsonify({"message": "Attendance not found"}), 404

    db.session.delete(row)
    db.session.commit()
    return jsonify({"message": "Attendance deleted successfully"}), 200


# -----------------------------
# 5) Import Attendance (CSV/XLSX) with UPSERT
# -----------------------------
@attendance_bp.route("/import", methods=["POST"])
@token_required
@role_required(ALLOWED_MANAGE_ROLES)
def import_attendance():
    """
    Supports:
      - CSV
      - XLSX (optional)
    Columns accepted:
      employee_id (required) OR employee_code (if you store it in Employee.employee_id string)
      date (required)
      status (Present/Absent)
      login_at (HH:MM)
      logout_at (HH:MM)

    UPSERT:
      company_id + employee_id + date exists => update
    """
    if "file" not in request.files:
        return jsonify({"message": "file is required (multipart/form-data)"}), 400

    f = request.files["file"]
    filename = (f.filename or "").lower()

    inserted = 0
    updated = 0
    errors = []

    company_id = g.user.company_id

    def handle_row(raw, row_no):
        nonlocal inserted, updated, errors

        try:
            # employee_id required
            emp_id = raw.get("employee_id")
            if not emp_id:
                raise ValueError("employee_id is required")

            emp = Employee.query.filter_by(employee_id=emp_id, company_id=company_id).first()
            if not emp:
                raise ValueError(f"Employee not found or not in company (employee_id={emp_id})")

            d = _parse_date(raw.get("date", ""))

            status = raw.get("status", "Present") or "Present"

            login_at = _parse_time(raw.get("login_at", ""), d) if raw.get("login_at") else None
            logout_at = _parse_time(raw.get("logout_at", ""), d) if raw.get("logout_at") else None

            payload = {"status": status, "login_at": login_at, "logout_at": logout_at}

            row, is_new = _upsert_attendance(company_id, emp.id, d, payload, capture_method="Import")
            if is_new:
                inserted += 1
            else:
                updated += 1

        except Exception as e:
            errors.append({"row": row_no, "error": str(e), "data": raw})

    # CSV
    if filename.endswith(".csv"):
        stream = io.StringIO(f.stream.read().decode("utf-8", errors="ignore"))
        reader = csv.DictReader(stream)

        for idx, r in enumerate(reader, start=2):  # 2 => header is row 1
            # normalize keys
            row = { (k or "").strip().lower(): (v or "").strip() for k, v in r.items() }
            handle_row(row, idx)

        db.session.commit()
        return jsonify({
            "message": "Import completed",
            "inserted": inserted,
            "updated": updated,
            "errors_count": len(errors),
            "errors": errors[:50]  # return first 50 only
        }), 200

    # XLSX (optional)
    if filename.endswith(".xlsx"):
        try:
            import openpyxl
        except ImportError:
            return jsonify({"message": "openpyxl not installed. Install it or use CSV."}), 400

        wb = openpyxl.load_workbook(f, data_only=True)
        ws = wb.active

        headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]

        for i in range(2, ws.max_row + 1):
            row_obj = {}
            for j, h in enumerate(headers, start=1):
                val = ws.cell(row=i, column=j).value
                row_obj[h] = str(val).strip() if val is not None else ""

            handle_row(row_obj, i)

        db.session.commit()
        return jsonify({
            "message": "Import completed",
            "inserted": inserted,
            "updated": updated,
            "errors_count": len(errors),
            "errors": errors[:50]
        }), 200

    return jsonify({"message": "Unsupported file type. Use .csv or .xlsx"}), 400


# -----------------------------
# 6) Employee view only (No create)
# -----------------------------
@attendance_bp.route("/me", methods=["GET"])
@token_required
@role_required(["EMPLOYEE", "ADMIN", "HR", "MANAGER", "SUPER_ADMIN"])
def my_attendance():
    """
    Employee can only view their own attendance.
    No punch in/out APIs here.
    """
    query = Employee.query.filter_by(user_id=g.user.id)
    if g.user.role != 'SUPER_ADMIN':
        query = query.filter_by(company_id=g.user.company_id)
    emp = query.first()
    if not emp:
        return jsonify({"message": "Employee profile not found"}), 404

    q = Attendance.query.filter_by(employee_id=emp.id)\
                        .order_by(Attendance.attendance_date.desc())\
                        .limit(180)

    output = []
    for r in q.all():
        output.append({
            "status": r.status,
            "logged_time": _format_logged_time(r.total_minutes),
            "login_at": r.punch_in_time.strftime("%H:%M") if r.punch_in_time else None,
            "logout_at": r.punch_out_time.strftime("%H:%M") if r.punch_out_time else None,
            "date": r.attendance_date.strftime("%d/%m/%Y"),
        })

    return jsonify({"attendance": output}), 200


# -----------------------------
# 7) Login / Logout APIs
# -----------------------------
@attendance_bp.route("/login", methods=["POST"])
@token_required
def attendance_login():
    data = request.get_json()
    employee_id = data.get("employee_id")
    status = data.get("status", "Present")

    if not employee_id:
        return jsonify({"message": "employee_id required"}), 400

    # Build query to find employee, handling Super Admin case
    query = Employee.query.filter_by(employee_id=employee_id)
    if g.user.role != 'SUPER_ADMIN':
        query = query.filter_by(company_id=g.user.company_id)
    emp = query.first()

    if not emp:
        return jsonify({"message": "Employee not found"}), 404

    today = date.today()

    attendance = Attendance.query.filter_by(
        employee_id=emp.id,
        attendance_date=today
    ).first()

    if attendance:
        return jsonify({"message": "Already logged in"}), 400

    punch_in_time = datetime.now()
    if status == "Absent":
        punch_in_time = None

    attendance = Attendance(
        employee_id=emp.id,
        company_id=emp.company_id,
        attendance_date=today,
        punch_in_time=punch_in_time,
        status=status,
        capture_method="Manual",
        created_by=g.user.id
    )

    db.session.add(attendance)
    db.session.commit()

    return jsonify({"message": "Login recorded"}), 201

@attendance_bp.route("/logout", methods=["POST"])
@token_required
def attendance_logout():
    data = request.get_json()
    employee_id = data.get("employee_id")

    # Build query to find employee, handling Super Admin case
    query = Employee.query.filter_by(employee_id=employee_id)
    if g.user.role != 'SUPER_ADMIN':
        query = query.filter_by(company_id=g.user.company_id)
    emp = query.first()
    if not emp: return jsonify({"message": "Employee not found"}), 404

    today = date.today()
    attendance = Attendance.query.filter_by(employee_id=emp.id, attendance_date=today).first()

    if not attendance: return jsonify({"message": "Login not found"}), 404
    if attendance.punch_out_time: return jsonify({"message": "Already logged out"}), 400

    attendance.punch_out_time = datetime.now()
    attendance.updated_by = g.user.id
    attendance.recalc_total_minutes()
    db.session.commit()

    return jsonify({"message": "Logout recorded"}), 200


# -----------------------------
# 8) Regularization (Correction)
# -----------------------------
@attendance_bp.route("/regularization/request", methods=["POST"])
@token_required
def create_regularization_request():
    data = request.get_json()
    
    # 1. Identify Employee
    emp = Employee.query.filter_by(user_id=g.user.id).first()
    if not emp:
        return jsonify({"message": "Employee profile not found"}), 404

    # 2. Validate Inputs
    att_date_str = data.get("attendance_date")
    if not att_date_str:
        return jsonify({"message": "attendance_date is required"}), 400
    
    try:
        d = _parse_date(att_date_str)
    except ValueError as e:
        return jsonify({"message": str(e)}), 400

    # 3. Parse Times (if provided)
    login_at = None
    logout_at = None
    if data.get("requested_login_at"):
        login_at = _parse_time(data["requested_login_at"], d)
    if data.get("requested_logout_at"):
        logout_at = _parse_time(data["requested_logout_at"], d)

    # 4. Create Request
    reg = AttendanceRegularization(
        company_id=emp.company_id,
        employee_id=emp.id,
        attendance_date=d,
        requested_status=data.get("requested_status"),
        requested_punch_in=login_at,
        requested_punch_out=logout_at,
        reason=data.get("reason"),
        status="PENDING"
    )
    
    db.session.add(reg)
    db.session.commit()

    return jsonify({
        "message": "Regularization request submitted",
        "request_id": reg.id,
        "status": "PENDING"
    }), 201

@attendance_bp.route("/regularization/my-requests", methods=["GET"])
@token_required
def my_regularization_requests():
    emp = Employee.query.filter_by(user_id=g.user.id).first()
    if not emp:
        return jsonify({"message": "Employee profile not found"}), 404

    reqs = AttendanceRegularization.query.filter_by(employee_id=emp.id).order_by(AttendanceRegularization.created_at.desc()).all()
    
    output = []
    for r in reqs:
        # Fetch current status
        att = Attendance.query.filter_by(employee_id=emp.id, attendance_date=r.attendance_date).first()
        current_status = att.status if att else "Absent"

        output.append({
            "request_id": r.id,
            "attendance_date": r.attendance_date.strftime("%Y-%m-%d"),
            "current_status": current_status,
            "requested_status": r.requested_status,
            "status": r.status,
            "reason": r.reason,
            "created_at": r.created_at.strftime("%Y-%m-%d %H:%M")
        })
    return jsonify({"requests": output}), 200

@attendance_bp.route("/regularization/pending", methods=["GET"])
@token_required
@role_required(ALLOWED_MANAGE_ROLES)
def pending_regularization_requests():
    q = AttendanceRegularization.query
    
    # Super Admin sees all, others see only their company
    if g.user.role != 'SUPER_ADMIN':
        q = q.filter_by(company_id=g.user.company_id)

    # Filters
    status = request.args.get("status", "PENDING")
    if status:
        q = q.filter_by(status=status)

    rows = q.order_by(AttendanceRegularization.created_at.desc()).all()
    
    output = []
    for r in rows:
        # Fetch current status
        att = Attendance.query.filter_by(employee_id=r.employee_id, attendance_date=r.attendance_date).first()
        current_status = att.status if att else "Absent"

        output.append({
            "request_id": r.id,
            "employee_id": r.employee.employee_id if r.employee else "",
            "employee_name": r.employee.full_name if r.employee else "Unknown",
            "attendance_date": r.attendance_date.strftime("%Y-%m-%d"),
            "current_status": current_status,
            "requested_status": r.requested_status,
            "requested_login_at": r.requested_punch_in.strftime("%H:%M") if r.requested_punch_in else None,
            "requested_logout_at": r.requested_punch_out.strftime("%H:%M") if r.requested_punch_out else None,
            "reason": r.reason,
            "status": r.status
        })
    return jsonify({"pending": output}), 200

@attendance_bp.route("/regularization/<int:request_id>/approve", methods=["POST"])
@token_required
@role_required(ALLOWED_MANAGE_ROLES)
def approve_regularization(request_id):
    data = request.get_json() or {}
    req = AttendanceRegularization.query.get_or_404(request_id)
    
    if g.user.role != 'SUPER_ADMIN' and req.company_id != g.user.company_id:
        return jsonify({"message": "Unauthorized"}), 403

    req.status = "APPROVED"
    req.approved_by = g.user.id
    req.approver_comment = data.get("approver_comment")

    # --- UPDATE ATTENDANCE ---
    att = Attendance.query.filter_by(employee_id=req.employee_id, attendance_date=req.attendance_date).first()
    if not att:
        att = Attendance(company_id=req.company_id, employee_id=req.employee_id, attendance_date=req.attendance_date, created_by=g.user.id)
        db.session.add(att)

    if req.requested_status: att.status = req.requested_status
    if req.requested_punch_in: att.punch_in_time = req.requested_punch_in
    if req.requested_punch_out: att.punch_out_time = req.requested_punch_out
    
    att.capture_method = "Regularization"
    att.updated_by = g.user.id
    att.recalc_total_minutes()
    
    db.session.commit()
    return jsonify({"message": "Request approved and attendance updated", "request_id": req.id, "attendance_action": "updated"}), 200

@attendance_bp.route("/regularization/<int:request_id>/reject", methods=["POST"])
@token_required
@role_required(ALLOWED_MANAGE_ROLES)
def reject_regularization(request_id):
    data = request.get_json() or {}
    req = AttendanceRegularization.query.get_or_404(request_id)

    if g.user.role != 'SUPER_ADMIN' and req.company_id != g.user.company_id:
        return jsonify({"message": "Unauthorized"}), 403

    req.status = "REJECTED"
    req.approved_by = g.user.id
    req.approver_comment = data.get("approver_comment")
    
    db.session.commit()
    return jsonify({"message": "Request rejected", "request_id": req.id, "status": "REJECTED"}), 200