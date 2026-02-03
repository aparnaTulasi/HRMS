import os
import shutil

def create_file(path, content):
    """Helper function to create a file with content."""
    dirname = os.path.dirname(path)
    if dirname:
        os.makedirs(dirname, exist_ok=True)
    with open(path, 'w', encoding='utf-8') as f:
        f.write(content.strip())
    print(f"✅ Created/Updated: {path}")

def delete_path(path):
    """Helper function to delete a file or directory."""
    if not os.path.exists(path):
        return
    try:
        if os.path.isdir(path):
            shutil.rmtree(path)
            print(f"🗑️  Deleted directory: {path}")
        else:
            os.remove(path)
            print(f"🗑️  Deleted file: {path}")
    except Exception as e:
        print(f"❌ Error deleting {path}: {e}")

# ==============================================================================
# CORRECT FILE CONTENTS
# ==============================================================================

app_py_content = """
from flask import Flask, jsonify, request
from flask_cors import CORS
from config import Config
from models import db
import logging

# Force-enable Werkzeug logging
logging.basicConfig(level=logging.INFO)
werkzeug_logger = logging.getLogger("werkzeug")
werkzeug_logger.setLevel(logging.INFO)
werkzeug_logger.disabled = False

app = Flask(__name__)
app.config.from_object(Config)

CORS(app)
db.init_app(app)

# Register Models (Ensure they are loaded for SQLAlchemy)
import models.user
import models.company
import models.employee
import models.permission
import models.attendance
import models.department
import models.filter
import models.urls
import models.payroll
import models.employee_address
import models.employee_bank
import models.employee_documents
import leave.models

# Import blueprints
from routes.auth import auth_bp
from routes.admin import admin_bp
from routes.employee import employee_bp
from routes.superadmin import superadmin_bp
from routes.hr import hr_bp
from routes.attendance import attendance_bp
from routes.employee_advanced import employee_advanced_bp
from routes.urls import urls_bp
from routes.permissions import permissions_bp
from routes.documents import documents_bp
from leave.routes import leave_bp

# Register blueprints
app.register_blueprint(auth_bp, url_prefix='/api/auth')
app.register_blueprint(admin_bp, url_prefix='/api/admin')
app.register_blueprint(superadmin_bp, url_prefix='/api/superadmin')
app.register_blueprint(hr_bp, url_prefix='/api/hr')
app.register_blueprint(employee_bp, url_prefix='/api/employee')
app.register_blueprint(attendance_bp, url_prefix='/api/attendance')
app.register_blueprint(employee_advanced_bp, url_prefix='/api/employee')
app.register_blueprint(urls_bp, url_prefix='/api/urls')
app.register_blueprint(permissions_bp, url_prefix='/api/permissions')
app.register_blueprint(documents_bp, url_prefix='/api/documents')
app.register_blueprint(leave_bp)

@app.before_request
def log_request():
    print(f"➡️ {request.method} {request.path}", flush=True)

@app.route('/')
def home():
    return jsonify({'message': 'HRMS API Running', 'version': '2.0'})

if __name__ == '__main__':
    with app.app_context():
        db.create_all()
    print("✅ HRMS Server Starting...")
    app.run(debug=True, port=5000)
"""

config_py_content = """
import os
from dotenv import load_dotenv

load_dotenv()

BASE_DIR = os.path.abspath(os.path.dirname(__file__))
INSTANCE_DIR = os.path.join(BASE_DIR, "instance")
os.makedirs(INSTANCE_DIR, exist_ok=True)

class Config:
    SECRET_KEY = os.getenv("SECRET_KEY", "hrms-secret-key-change-this")
    SQLALCHEMY_DATABASE_URI = f"sqlite:///{os.path.join(INSTANCE_DIR, 'hrms.db')}"
    SQLALCHEMY_TRACK_MODIFICATIONS = False
    UPLOAD_FOLDER = os.path.join(BASE_DIR, "uploads")
    BASE_URL = "http://localhost:5000"

    # Email Configuration
    MAIL_SERVER = "smtp.gmail.com"
    MAIL_USE_TLS = True
    MAIL_USERNAME = os.getenv("MAIL_USERNAME")
    MAIL_PASSWORD = os.getenv("MAIL_PASSWORD")
    MAIL_PORT = 587
    MAIL_DEFAULT_SENDER = os.getenv("MAIL_DEFAULT_SENDER")

os.makedirs(Config.UPLOAD_FOLDER, exist_ok=True)
"""

models_init_py_content = "from flask_sqlalchemy import SQLAlchemy\ndb = SQLAlchemy()"

models_user_py_content = """
from datetime import datetime, timedelta
from flask_login import UserMixin
from models import db
import secrets

class User(UserMixin, db.Model):
    __tablename__ = 'users'
    id = db.Column(db.Integer, primary_key=True)
    email = db.Column(db.String(120), unique=True, nullable=False)
    password = db.Column(db.String(255), nullable=False)
    role = db.Column(db.String(20), nullable=False)
    company_id = db.Column(db.Integer, db.ForeignKey('companies.id'), nullable=True)
    status = db.Column(db.String(20), default='ACTIVE')
    portal_prefix = db.Column(db.String(50), nullable=True)
    otp = db.Column(db.String(6), nullable=True)
    otp_expiry = db.Column(db.DateTime, nullable=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)
    employee_profile = db.relationship('Employee', backref='user', uselist=False, lazy=True)
    permissions = db.relationship('UserPermission', foreign_keys='UserPermission.user_id', backref=db.backref('user', foreign_keys='UserPermission.user_id'), lazy=True)

    def generate_otp(self):
        self.otp = ''.join(secrets.choice('0123456789') for _ in range(6))
        self.otp_expiry = datetime.utcnow() + timedelta(minutes=10)
        return self.otp

    def has_permission(self, permission_code):
        if self.role == 'SUPER_ADMIN':
            return True
        return any(p.permission_code == permission_code for p in self.permissions)
"""

models_company_py_content = """
from datetime import datetime
from models import db

class Company(db.Model):
    __tablename__ = 'companies'
    id = db.Column(db.Integer, primary_key=True)
    company_name = db.Column(db.String(100), nullable=False)
    subdomain = db.Column(db.String(50), unique=True, nullable=False)
    company_code = db.Column(db.String(20), unique=True)
    industry = db.Column(db.String(100))
    company_size = db.Column(db.String(50))
    state = db.Column(db.String(100))
    country = db.Column(db.String(100))
    city_branch = db.Column(db.String(100))
    timezone = db.Column(db.String(50))
    address = db.Column(db.Text)
    phone = db.Column(db.String(20))
    email = db.Column(db.String(120))
    website = db.Column(db.String(200))
    attendance_enabled = db.Column(db.Boolean, default=True)
    leave_enabled = db.Column(db.Boolean, default=True)
    payroll_enabled = db.Column(db.Boolean, default=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)
    users = db.relationship('User', backref='company', lazy=True)
    employees = db.relationship('Employee', backref='company', lazy=True)
    departments = db.relationship('Department', backref='company', lazy=True)
"""

models_employee_py_content = """
from datetime import datetime
from models import db

class Employee(db.Model):
    __tablename__ = 'employees'
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('users.id'), unique=True, nullable=False)
    company_id = db.Column(db.Integer, db.ForeignKey('companies.id'), nullable=True)
    company_code = db.Column(db.String(20))
    employee_id = db.Column(db.String(50), unique=True)
    first_name = db.Column(db.String(50), nullable=False)
    last_name = db.Column(db.String(50), nullable=False)
    gender = db.Column(db.String(10))
    date_of_birth = db.Column(db.Date)
    father_or_husband_name = db.Column(db.String(100))
    mother_name = db.Column(db.String(100))
    department = db.Column(db.String(50))
    designation = db.Column(db.String(50))
    salary = db.Column(db.Float(10,2))
    date_of_joining = db.Column(db.Date)
    work_phone = db.Column(db.String(20))
    personal_mobile = db.Column(db.String(20))
    personal_email = db.Column(db.String(120))
    company_email = db.Column(db.String(120))
    work_mode = db.Column(db.String(20))
    branch_id = db.Column(db.Integer)
    aadhaar_number = db.Column(db.String(20), unique=True)
    pan_number = db.Column(db.String(20), unique=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)
    bank_details = db.relationship('EmployeeBankDetails', backref='employee', uselist=False, lazy=True)
    addresses = db.relationship('EmployeeAddress', backref='employee', lazy=True)
    documents = db.relationship('EmployeeDocument', backref='employee', lazy=True)
    attendance_records = db.relationship('Attendance', backref='employee', lazy=True)


    @property
    def full_name(self):
        return f"{self.first_name} {self.last_name}"
"""

models_attendance_py_content = '''
from datetime import datetime, date
from models import db

class Attendance(db.Model):
    """
    One row per employee per date (UPSERT key)
    """
    __tablename__ = "attendance_logs"

    attendance_id = db.Column(db.Integer, primary_key=True)

    company_id = db.Column(db.Integer, db.ForeignKey("companies.id"), nullable=False, index=True)
    employee_id = db.Column(db.Integer, db.ForeignKey("employees.id"), nullable=False, index=True)

    # Main unique key for UPSERT
    attendance_date = db.Column(db.Date, nullable=False, index=True)

    # Times (optional for Absent)
    punch_in_time = db.Column(db.DateTime, nullable=True)
    punch_out_time = db.Column(db.DateTime, nullable=True)

    # Stored to show "Logged Time"
    total_minutes = db.Column(db.Integer, default=0, nullable=False)

    # Present / Absent / Leave / Half Day etc. (keep it flexible)
    status = db.Column(db.String(20), default="Present", nullable=False)

    # manual/import/web/biometric (you said no self punch, still keep for audit)
    capture_method = db.Column(db.String(50), default="Manual", nullable=False)  # Manual / Import

    # Audit
    created_by = db.Column(db.Integer, db.ForeignKey("users.id"), nullable=True)
    updated_by = db.Column(db.Integer, db.ForeignKey("users.id"), nullable=True)

    created_at = db.Column(db.DateTime, default=datetime.utcnow, nullable=False)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow, nullable=False)

    __table_args__ = (
        db.UniqueConstraint("company_id", "employee_id", "attendance_date", name="uq_att_company_emp_date"),
    )

    def recalc_total_minutes(self):
        if self.punch_in_time and self.punch_out_time and self.punch_out_time > self.punch_in_time:
            diff = self.punch_out_time - self.punch_in_time
            self.total_minutes = int(diff.total_seconds() // 60)
        else:
            self.total_minutes = 0
'''

models_permission_py_content = """
from datetime import datetime
from models import db

class Permission(db.Model):
    __tablename__ = 'permissions'
    id = db.Column(db.Integer, primary_key=True)
    permission_code = db.Column(db.String(50), unique=True, nullable=False)
    description = db.Column(db.String(255))
    module = db.Column(db.String(50))
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

class UserPermission(db.Model):
    __tablename__ = 'user_permissions'
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('users.id'), nullable=False)
    permission_code = db.Column(db.String(50), nullable=False)
    granted_at = db.Column(db.DateTime, default=datetime.utcnow)
    granted_by = db.Column(db.Integer, db.ForeignKey('users.id'))
"""

models_department_py_content = """
from datetime import datetime
from models import db

class Department(db.Model):
    __tablename__ = 'departments'
    id = db.Column(db.Integer, primary_key=True)
    company_id = db.Column(db.Integer, db.ForeignKey('companies.id'), nullable=False)
    department_name = db.Column(db.String(100), nullable=False)
    department_code = db.Column(db.String(20), unique=True)
    description = db.Column(db.Text)
    manager_id = db.Column(db.Integer, db.ForeignKey('employees.id'))
    is_active = db.Column(db.Boolean, default=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)
    manager = db.relationship('Employee', foreign_keys=[manager_id])
"""

models_filter_py_content = """
from datetime import datetime
from models import db
import json

class FilterConfiguration(db.Model):
    __tablename__ = 'filter_configurations'
    id = db.Column(db.Integer, primary_key=True)
    filter_name = db.Column(db.String(100), nullable=False)
    module = db.Column(db.String(50), nullable=False)
    filter_config = db.Column(db.Text)
    allowed_roles = db.Column(db.Text)
    company_id = db.Column(db.Integer, db.ForeignKey('companies.id'), nullable=True)
    created_by = db.Column(db.Integer, db.ForeignKey('users.id'))
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)
"""

models_urls_py_content = """
from datetime import datetime
from models import db

class SystemURL(db.Model):
    __tablename__ = 'system_urls'
    id = db.Column(db.Integer, primary_key=True)
    url_code = db.Column(db.String(50), unique=True, nullable=False)
    url_path = db.Column(db.String(255), nullable=False)
    description = db.Column(db.String(255))
    module = db.Column(db.String(50))
    allowed_roles = db.Column(db.Text)
    permission_required = db.Column(db.String(50))
    is_active = db.Column(db.Boolean, default=True)
    is_public = db.Column(db.Boolean, default=False)
    company_id = db.Column(db.Integer, db.ForeignKey('companies.id'), nullable=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)
    company = db.relationship('Company', backref='urls')
"""

models_payroll_py_content = """
from datetime import datetime
from models import db
from sqlalchemy import CheckConstraint

class SalaryComponent(db.Model):
    __tablename__ = 'salary_components'
    id = db.Column(db.Integer, primary_key=True) # component_id
    company_id = db.Column(db.Integer, db.ForeignKey('companies.id'), nullable=False)
    
    component_name = db.Column(db.String(100), nullable=False)
    component_code = db.Column(db.String(20), nullable=False)
    type = db.Column(db.String(20), nullable=False) # Earning/Deduction
    calculation_type = db.Column(db.String(20), nullable=False) # Fixed/Percentage
    percentage_value = db.Column(db.Float(5,2))
    taxable = db.Column(db.String(5), default='Yes', nullable=False) # Yes/No
    order_no = db.Column(db.Integer, nullable=False)
    is_active = db.Column(db.String(5), default='Yes', nullable=False) # Yes/No

    __table_args__ = (
        db.UniqueConstraint('component_name', 'company_id', name='uq_component_name_company'),
        db.UniqueConstraint('component_code', 'company_id', name='uq_component_code_company'),
        CheckConstraint("type IN ('Earning', 'Deduction')", name='chk_comp_type'),
        CheckConstraint("calculation_type IN ('Fixed', 'Percentage')", name='chk_calc_type'),
        CheckConstraint("taxable IN ('Yes', 'No')", name='chk_taxable'),
        CheckConstraint("is_active IN ('Yes', 'No')", name='chk_active_comp'),
    )

class SalaryStructure(db.Model):
    __tablename__ = 'salary_structures'
    id = db.Column(db.Integer, primary_key=True) # structure_id
    company_id = db.Column(db.Integer, db.ForeignKey('companies.id'), nullable=False)
    
    structure_name = db.Column(db.String(100), nullable=False)
    description = db.Column(db.String(300))
    base_salary = db.Column(db.Float(10,2))
    is_active = db.Column(db.String(5), default='Yes', nullable=False)
    
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)

    __table_args__ = (
        db.UniqueConstraint('structure_name', 'company_id', name='uq_structure_name_company'),
        CheckConstraint("is_active IN ('Yes', 'No')", name='chk_active_struct'),
    )
    
    components = db.relationship('SalaryStructureComponent', backref='structure', lazy=True)

class SalaryStructureComponent(db.Model):
    __tablename__ = 'salary_structure_components'
    id = db.Column(db.Integer, primary_key=True)
    structure_id = db.Column(db.Integer, db.ForeignKey('salary_structures.id'), nullable=False)
    component_id = db.Column(db.Integer, db.ForeignKey('salary_components.id'), nullable=False)
    
    percentage = db.Column(db.Float(5,2))
    fixed_amount = db.Column(db.Float(10,2))
    depends_on_component_id = db.Column(db.Integer, db.ForeignKey('salary_components.id'), nullable=True)

    __table_args__ = (
        db.UniqueConstraint('structure_id', 'component_id', name='uq_ssc'),
    )
    
    component = db.relationship('SalaryComponent', foreign_keys=[component_id])
    depends_on = db.relationship('SalaryComponent', foreign_keys=[depends_on_component_id])

class EmployeeSalaryStructure(db.Model):
    __tablename__ = 'employee_salary_structure'
    id = db.Column(db.Integer, primary_key=True)
    employee_id = db.Column(db.Integer, db.ForeignKey('employees.id'), nullable=False)
    structure_id = db.Column(db.Integer, db.ForeignKey('salary_structures.id'), nullable=False)
    
    effective_from = db.Column(db.Date, nullable=False)
    effective_to = db.Column(db.Date)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)

    __table_args__ = (
        db.UniqueConstraint('employee_id', 'effective_from', name='uq_ess'),
    )

class PayrollRun(db.Model):
    __tablename__ = 'payroll_run'
    id = db.Column(db.Integer, primary_key=True) # payroll_id
    company_id = db.Column(db.Integer, db.ForeignKey('companies.id'), nullable=False)
    month_year = db.Column(db.String(20), nullable=False)
    run_date = db.Column(db.Date, default=datetime.utcnow)
    status = db.Column(db.String(20), default='In Progress', nullable=False) # In Progress, Completed, Locked
    created_by = db.Column(db.Integer, db.ForeignKey('users.id'))
    remarks = db.Column(db.String(300))

    __table_args__ = (
        db.UniqueConstraint('month_year', 'company_id', name='uq_pr_month_company'),
        CheckConstraint("status IN ('In Progress', 'Completed', 'Locked')", name='chk_payroll_status'),
    )

class PayrollRunEmployee(db.Model):
    __tablename__ = 'payroll_run_employees'
    id = db.Column(db.Integer, primary_key=True)
    payroll_id = db.Column(db.Integer, db.ForeignKey('payroll_run.id'), nullable=False)
    employee_id = db.Column(db.Integer, db.ForeignKey('employees.id'), nullable=False)
    paid_days = db.Column(db.Float(5,2))
    lop_days = db.Column(db.Float(5,2))
    overtime_hours = db.Column(db.Float(5,2))
    overtime_amount = db.Column(db.Float(10,2))

    __table_args__ = (
        db.UniqueConstraint('payroll_id', 'employee_id', name='uq_pre'),
    )

class PayrollEarning(db.Model):
    __tablename__ = 'payroll_earnings'
    id = db.Column(db.Integer, primary_key=True) # earning_id
    payroll_id = db.Column(db.Integer, db.ForeignKey('payroll_run.id'), nullable=False)
    employee_id = db.Column(db.Integer, db.ForeignKey('employees.id'), nullable=False)
    component_id = db.Column(db.Integer, db.ForeignKey('salary_components.id'), nullable=False)
    amount = db.Column(db.Float(10,2), nullable=False)

    __table_args__ = (
        db.UniqueConstraint('payroll_id', 'employee_id', 'component_id', name='uq_pe'),
    )

class PayrollDeduction(db.Model):
    __tablename__ = 'payroll_deductions'
    id = db.Column(db.Integer, primary_key=True) # deduction_id
    payroll_id = db.Column(db.Integer, db.ForeignKey('payroll_run.id'), nullable=False)
    employee_id = db.Column(db.Integer, db.ForeignKey('employees.id'), nullable=False)
    component_id = db.Column(db.Integer, db.ForeignKey('salary_components.id'), nullable=False)
    amount = db.Column(db.Float(10,2), nullable=False)
    is_manual = db.Column(db.String(5), default='No') # Yes/No

    __table_args__ = (
        db.UniqueConstraint('payroll_id', 'employee_id', 'component_id', name='uq_pd'),
        CheckConstraint("is_manual IN ('Yes', 'No')", name='chk_is_manual'),
    )

class PayrollSummary(db.Model):
    __tablename__ = 'payroll_summary'
    id = db.Column(db.Integer, primary_key=True) # summary_id
    payroll_id = db.Column(db.Integer, db.ForeignKey('payroll_run.id'), nullable=False)
    employee_id = db.Column(db.Integer, db.ForeignKey('employees.id'), nullable=False)
    gross_salary = db.Column(db.Float(12,2))
    total_deductions = db.Column(db.Float(12,2))
    net_salary = db.Column(db.Float(12,2))
    employer_contribution = db.Column(db.Float(12,2))

    __table_args__ = (
        db.UniqueConstraint('payroll_id', 'employee_id', name='uq_ps'),
    )
"""

models_employee_address_py_content = """
from models import db
from datetime import datetime

class EmployeeAddress(db.Model):
    __tablename__ = 'employee_address'
    id = db.Column(db.Integer, primary_key=True)
    employee_id = db.Column(db.Integer, db.ForeignKey('employees.id'), nullable=False)
    address_type = db.Column(db.String(20)) # PRESENT / PERMANENT
    address_line1 = db.Column(db.String(150))
    address_line2 = db.Column(db.String(150))
    city = db.Column(db.String(100))
    state = db.Column(db.String(100))
    zip_code = db.Column(db.String(10))
    country = db.Column(db.String(100))
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)
"""

models_employee_bank_py_content = """
from models import db
from datetime import datetime

class EmployeeBankDetails(db.Model):
    __tablename__ = 'employee_bank_details'
    id = db.Column(db.Integer, primary_key=True)
    employee_id = db.Column(db.Integer, db.ForeignKey('employees.id'), unique=True, nullable=False)
    bank_name = db.Column(db.String(100), nullable=False)
    branch_name= db.Column(db.String(100),nullable=False)
    account_number = db.Column(db.String(20), unique=True, nullable=False)
    ifsc_code = db.Column(db.String(20), nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)
"""

models_employee_documents_py_content = """
from datetime import datetime
from models import db

class EmployeeDocument(db.Model):
    __tablename__ = 'employee_documents'
    id = db.Column(db.Integer, primary_key=True)
    employee_id = db.Column(db.Integer, db.ForeignKey('employees.id'), nullable=False)
    document_type = db.Column(db.String(100), nullable=False)
    document_number = db.Column(db.String(100))
    document_name = db.Column(db.String(255))
    file_path = db.Column(db.String(500))
    file_url = db.Column(db.String(500))
    issue_date = db.Column(db.Date)
    expiry_date = db.Column(db.Date)
    verified = db.Column(db.Boolean, default=False)
    verified_by = db.Column(db.Integer, db.ForeignKey('users.id'))
    verified_date = db.Column(db.DateTime)
    notes = db.Column(db.Text)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)

class DocumentType(db.Model):
    __tablename__ = 'document_types'
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False, unique=True)
    description = db.Column(db.Text)
    is_required = db.Column(db.Boolean, default=False)
    requires_expiry_date = db.Column(db.Boolean, default=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
"""

routes_auth_py_content = """
from flask import Blueprint, request, jsonify
from werkzeug.security import generate_password_hash, check_password_hash
import jwt
import datetime
import secrets
from models import db
from models.user import User
from models.company import Company
from models.employee import Employee
from config import Config
from utils.email_utils import send_login_success_email, send_signup_otp
from utils.url_generator import build_company_base_url

auth_bp = Blueprint('auth', __name__)

@auth_bp.route('/register', methods=['POST'])
def register():
    data = request.get_json()
    if User.query.filter_by(email=data['email']).first():
        return jsonify({'message': 'User already exists'}), 409

    hashed_password = generate_password_hash(data['password'], method='pbkdf2:sha256')
    new_user = User(email=data['email'], password=hashed_password, role=data.get('role', 'EMPLOYEE'))
    db.session.add(new_user)
    db.session.commit()
    return jsonify({'message': 'User registered successfully'}), 201

@auth_bp.route('/login', methods=['POST'])
def login():
    data = request.get_json()
    email = data.get('email') or data.get('company_email')
    if not email:
        return jsonify({'message': 'Email is required'}), 400

    user = User.query.filter_by(email=email).first()
    if not user or not check_password_hash(user.password, data['password']):
        return jsonify({'message': 'Invalid credentials'}), 401

    token = jwt.encode({
        'user_id': user.id, 'email': user.email, 'role': user.role,
        'exp': datetime.datetime.utcnow() + datetime.timedelta(hours=24)
    }, Config.SECRET_KEY, algorithm="HS256")
    
    try:
        send_login_success_email(user.email)
    except:
        pass

    # Determine Redirect URL
    ROLE_PATHS = {
        "SUPER_ADMIN": "/super/dashboard",
        "ADMIN": "/admin/dashboard",
        "HR": "/hr/dashboard",
        "EMPLOYEE": "/employee/dashboard",
    }
    
    company = Company.query.get(user.company_id) if user.company_id else None
    subdomain = company.subdomain if company else ""
    base_url = build_company_base_url(subdomain)
    path = ROLE_PATHS.get(user.role, "/")
    
    return jsonify({
        'token': token,
        'role': user.role,
        'redirect_url': f"{base_url}{path}"
    })

@auth_bp.route('/super-admin/signup', methods=['POST'])
def super_admin_signup():
    data = request.get_json()
    if User.query.filter_by(email=data['email']).first():
        return jsonify({'message': 'User already exists'}), 409

    hashed_password = generate_password_hash(data['password'], method='pbkdf2:sha256')
    new_user = User(email=data['email'], password=hashed_password, role='SUPER_ADMIN', status='PENDING_OTP')
    
    # Generate OTP
    otp = new_user.generate_otp()
    
    send_signup_otp(data['email'], otp)
    db.session.add(new_user)
    db.session.flush()

    new_employee = Employee(
        user_id=new_user.id,
        first_name=data.get('first_name'),
        last_name=data.get('last_name'),
        employee_id=f"SA-{new_user.id}"
    )
    db.session.add(new_employee)
    db.session.commit()

    return jsonify({'message': 'Super Admin registered successfully. Please check your email for OTP.'}), 201

@auth_bp.route('/verify-signup-otp', methods=['POST'])
def verify_signup_otp():
    data = request.get_json()
    user = User.query.filter_by(email=data.get('email')).first()
    
    if not user:
        return jsonify({'message': 'User not found'}), 404
    if user.otp != data.get('otp'):
        return jsonify({'message': 'Invalid OTP'}), 400
    if user.otp_expiry and user.otp_expiry < datetime.datetime.utcnow():
        return jsonify({'message': 'OTP expired'}), 400
        
    user.status = 'ACTIVE'
    user.otp = None
    user.otp_expiry = None
    db.session.commit()
    return jsonify({'message': 'Account verified successfully'}), 200
"""

routes_admin_py_content = """
from flask import Blueprint, jsonify, request, g
from werkzeug.security import generate_password_hash
from datetime import datetime
from models import db
from models.user import User
from models.company import Company
from utils.decorators import token_required, role_required
from models.employee import Employee
from utils.email_utils import send_account_created_alert, send_login_credentials
from utils.url_generator import build_web_address, build_common_login_url

admin_bp = Blueprint('admin', __name__)

@admin_bp.route('/employees', methods=['GET'])
@token_required
@role_required(['ADMIN', 'HR'])
def get_employees():
    employees = Employee.query.filter_by(company_id=g.user.company_id).all()
    return jsonify([{'id': emp.id, 'name': emp.full_name} for emp in employees])

@admin_bp.route('/create-hr', methods=['POST'])
@token_required
@role_required(['ADMIN'])
def create_hr():
    print("🔵 Create HR Request Received...", flush=True)
    data = request.get_json(force=True)
    email = data.get('email') or data.get('company_email')
    personal_email = data.get('personal_email')
    if not email or not data.get('password'):
        return jsonify({'message': 'Email and Password are required'}), 400

    if User.query.filter_by(email=email).first():
        return jsonify({'message': 'Email already exists'}), 409

    hashed_password = generate_password_hash(data['password'])
    new_user = User(email=email, password=hashed_password, role='HR', company_id=g.user.company_id)
    db.session.add(new_user)
    db.session.flush()

    company = Company.query.get(g.user.company_id)
    emp_count = Employee.query.filter_by(company_id=g.user.company_id).count()
    emp_code = f"{company.company_code}-{emp_count + 1:04d}"

    new_employee = Employee(
        user_id=new_user.id,
        company_id=g.user.company_id,
        company_code=company.company_code,
        employee_id=emp_code,
        first_name=data.get('first_name'),
        last_name=data.get('last_name'),
        company_email=email,
        personal_email=personal_email,
        department=data.get('department', 'Human Resources'),
        designation=data.get('designation', 'HR Manager')
    )
    db.session.add(new_employee)
    db.session.commit()

    # --- Terminal Logging & Email Sending ---
    print(f"✅ HR Created: {email} | Personal Email: {personal_email}", flush=True)

    if personal_email:
        try:
            company = Company.query.get(g.user.company_id)
            creator_name = g.user.employee_profile.full_name if g.user.employee_profile else "Admin"
            web_address = build_web_address(company.subdomain)
            login_url = build_common_login_url(company.subdomain)
            
            send_account_created_alert(personal_email, company.company_name, creator_name)
            send_login_credentials(
                personal_email=personal_email,
                company_email=email,
                password=data['password'],
                company_name=company.company_name,
                web_address=web_address,
                login_url=login_url,
                created_by=creator_name
            )
            print(f"📧 Credentials sent to {personal_email}", flush=True)
        except Exception as e:
            print(f"❌ Failed to send email: {e}", flush=True)

    return jsonify({'message': 'HR created successfully'}), 201

@admin_bp.route('/employees', methods=['POST'])
@admin_bp.route('/create-employee', methods=['POST'])
@token_required
@role_required(['ADMIN', 'HR'])
def create_employee():
    print("🔵 Create Employee Request Received...", flush=True)
    data = request.get_json(force=True)
    
    email = data.get('email') or data.get('company_email')
    required = ['first_name', 'last_name', 'password', 'personal_email']
    if not email or not all(k in data for k in required):
        return jsonify({'message': 'Missing required fields'}), 400
        
    company_id = g.user.company_id
    email = email.lower().strip()
    personal_email = data['personal_email'].lower().strip()
    
    if User.query.filter_by(email=email).first():
        return jsonify({'message': 'Email already exists'}), 409
        
    hashed_password = generate_password_hash(data['password'])
    new_user = User(
        email=email,
        password=hashed_password,
        role=data.get('role', 'EMPLOYEE'),
        company_id=company_id,
        status='ACTIVE'
    )
    db.session.add(new_user)
    db.session.flush()
    
    company = Company.query.get(company_id)
    emp_count = Employee.query.filter_by(company_id=company_id).count()
    emp_code = f"{company.company_code}-{emp_count + 1:04d}"
    
    date_of_joining = data.get('date_of_joining')
    if date_of_joining:
        try:
            date_of_joining = datetime.strptime(date_of_joining, '%Y-%m-%d').date()
        except ValueError:
            return jsonify({'message': 'Invalid date format. Use YYYY-MM-DD'}), 400

    new_employee = Employee(
        user_id=new_user.id,
        company_id=company_id,
        company_code=company.company_code,
        employee_id=emp_code,
        first_name=data['first_name'],
        last_name=data['last_name'],
        company_email=email,
        personal_email=personal_email,
        department=data.get('department'),
        designation=data.get('designation'),
        date_of_joining=date_of_joining
    )
    db.session.add(new_employee)
    db.session.commit()
    
    print(f"✅ Employee Created: {email} | Personal Email: {personal_email}", flush=True)

    try:
        # Send Emails
        creator_name = g.user.employee_profile.full_name if g.user.employee_profile else "Admin"
        web_address = build_web_address(company.subdomain)
        login_url = build_common_login_url(company.subdomain)
        
        # Mail 1: Alert
        send_account_created_alert(personal_email, company.company_name, creator_name)
        
        # Mail 2: Credentials
        send_login_credentials(
            personal_email=personal_email,
            company_email=email,
            password=data['password'],
            company_name=company.company_name,
            web_address=web_address,
            login_url=login_url,
            created_by=creator_name
        )
        print(f"📧 Credentials sent to {personal_email}", flush=True)
    except Exception as e:
        print(f"❌ Failed to send email: {e}", flush=True)
    
    return jsonify({'message': 'Employee created successfully'}), 201
"""

routes_superadmin_py_content = """
from flask import Blueprint, request, jsonify
from werkzeug.security import generate_password_hash
from datetime import datetime
from models import db
from models.company import Company
from models.user import User
from models.employee import Employee
from utils.decorators import token_required, role_required
from utils.email_utils import send_account_created_alert, send_login_credentials
from utils.url_generator import build_web_address, build_common_login_url

superadmin_bp = Blueprint('superadmin', __name__)

@superadmin_bp.route('/create-company', methods=['POST'])
@token_required
@role_required(['SUPER_ADMIN'])
def create_company():
    print("🔵 Create Company Request Received...", flush=True)
    data = request.get_json()
    
    if Company.query.filter_by(subdomain=data['subdomain']).first():
        return jsonify({'message': 'Subdomain already exists'}), 409

      new_company = Company(
        company_name=data['company_name'], 
        subdomain=data['subdomain'],
        company_code=data.get('company_code'),
        industry=data.get('industry')
    )
    allowed_fields = [
        "company_name", "subdomain", "company_code",
        "industry", "company_size",
        "state", "country", "city_branch", "timezone"
    ]
    filtered_data = {k: v for k, v in data.items() if k in allowed_fields}

    new_company = Company(**filtered_data)
    db.session.add(new_company)
    db.session.commit()
    
    return jsonify({'message': 'Company created successfully', 'company': {'id': new_company.id, 'name': new_company.company_name}}), 201

@superadmin_bp.route('/create-admin', methods=['POST'])
@token_required
@role_required(['SUPER_ADMIN'])
def create_admin():
    print("🔵 Create Admin Request Received...", flush=True)
    data = request.get_json()
    
    required_fields = ['company_id', 'company_email', 'password', 'first_name', 'last_name']
    if not all(k in data for k in required_fields):
        return jsonify({'message': 'Missing required fields'}), 400

    company = Company.query.get(data['company_id'])
    if not company:
        return jsonify({'message': 'Company not found'}), 404

    if User.query.filter_by(email=data['company_email']).first():
        return jsonify({'message': 'Email already exists'}), 409

    hashed_password = generate_password_hash(data['password'], method='pbkdf2:sha256')
    new_user = User(
        email=data['company_email'],
        password=hashed_password,
        role='ADMIN',
        company_id=company.id,
        status='ACTIVE'
    )
    db.session.add(new_user)
    db.session.flush()
    
    emp_count = Employee.query.filter_by(company_id=company.id).count()
    emp_code = f"{company.company_code}-ADMIN-{emp_count + 1:02d}" if company.company_code else f"ADMIN-{emp_count + 1:02d}"

    new_employee = Employee(
        user_id=new_user.id,
        company_id=company.id,
        company_code=company.company_code,
        employee_id=emp_code,
        first_name=data['first_name'],
        last_name=data['last_name'],
        company_email=data['company_email'],
        personal_email=data.get('personal_email'),
        department=data.get('department', 'Administration'),
        designation=data.get('designation', 'Company Admin'),
        date_of_joining=datetime.utcnow().date()
    )
    db.session.add(new_employee)
    db.session.commit()

    print(f"✅ Admin Created: {data['company_email']}", flush=True)

    if data.get('personal_email'):
        web_address = build_web_address(company.subdomain)
        login_url = build_common_login_url(company.subdomain)
        
        send_account_created_alert(data['personal_email'], company.company_name, "Super Admin")
        send_login_credentials(
            personal_email=data['personal_email'],
            company_email=data['company_email'],
            password=data['password'],
            company_name=company.company_name,
            web_address=web_address,
            login_url=login_url,
            created_by="Super Admin"
        )
        print(f"📧 Credentials sent to {data['personal_email']}", flush=True)

    return jsonify({'message': 'Company Admin created successfully'}), 201
"""


routes_hr_py_content = """
from flask import Blueprint, jsonify, request, g
from models import db
from models.user import User
from models.employee import Employee
from utils.decorators import token_required, role_required

hr_bp = Blueprint('hr', __name__)

@hr_bp.route('/employees', methods=['GET'])
@token_required
@role_required(['HR'])
def get_employees():
    employees = Employee.query.filter_by(company_id=g.user.company_id).all()
    output = []
    for emp in employees:
        user = User.query.get(emp.user_id)
        output.append({
            'id': emp.id,
            'first_name': emp.first_name,
            'last_name': emp.last_name,
            'email': user.email,
            'status': user.status,
            'department': emp.department,
            'designation': emp.designation
        })
    return jsonify({'employees': output})

@hr_bp.route('/approve-employee/<int:user_id>', methods=['POST'])
@token_required
@role_required(['HR'])
def approve_employee(user_id):
    user = User.query.get(user_id)
    if not user:
        return jsonify({'message': 'User not found'}), 404
    if user.company_id != g.user.company_id:
        return jsonify({'message': 'Unauthorized access'}), 403
    user.status = 'ACTIVE'
    db.session.commit()
    return jsonify({'message': 'Employee approved successfully', 'status': 'ACTIVE'})
"""

routes_employee_py_content = """
from flask import Blueprint, jsonify, request, g
from models import db
from models.employee import Employee
from models.employee_bank import EmployeeBankDetails
from models.employee_address import EmployeeAddress
from models.employee_documents import EmployeeDocument
from utils.decorators import token_required
import os
from config import Config
from werkzeug.utils import secure_filename

employee_bp = Blueprint('employee', __name__)

@employee_bp.route('/profile', methods=['GET'])
@token_required
def get_profile():
    emp = Employee.query.filter_by(user_id=g.user.id).first()
    if not emp:
        return jsonify({'message': 'Profile not found'}), 404
    
    addresses = []
    for addr in emp.addresses:
        addresses.append({
            'type': addr.address_type,
            'line1': addr.address_line1,
            'line2': addr.address_line2,
            'city': addr.city,
            'state': addr.state,
            'zip_code': addr.zip_code,
            'country': addr.country
        })

    return jsonify({
        'id': emp.id,
        'employee_id': emp.employee_id,
        'first_name': emp.first_name,
        'last_name': emp.last_name,
        'email': emp.company_email,
        'personal_email': emp.personal_email,
        'department': emp.department,
        'designation': emp.designation,
        'salary': emp.salary,
        'phone': getattr(emp, 'work_phone', None),
        'date_of_joining': emp.date_of_joining.isoformat() if emp.date_of_joining else None,
        'addresses': addresses
    })

# Other employee routes like /bank, /address etc. would go here
"""

routes_attendance_py_content = '''
from flask import Blueprint, request, jsonify, g
from datetime import datetime, date
import csv
import io

from models import db
from models.attendance import Attendance
from models.employee import Employee
from models.user import User
from utils.decorators import token_required, role_required

attendance_bp = Blueprint("attendance", __name__)

ALLOWED_MANAGE_ROLES = ["SUPER_ADMIN", "ADMIN", "HR"]


# -----------------------------
# Helpers
# -----------------------------
def _parse_date(value: str) -> date:
    """
    Accepts:
      - YYYY-MM-DD
      - DD/MM/YYYY
      - MM/DD/YYYY (if you want; can remove)
    """
    if not value:
        raise ValueError("date is required")

    value = value.strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(value, fmt).date()
        except ValueError:
            continue
    raise ValueError(f"Invalid date format: {value}")


def _parse_time(value: str, base_date: date) -> datetime:
    """
    Accepts:
      - HH:MM
      - HH:MM:SS
      - 09:45
      - 19:34
    """
    if not value:
        return None

    value = value.strip()
    for fmt in ("%H:%M", "%H:%M:%S"):
        try:
            t = datetime.strptime(value, fmt).time()
            return datetime.combine(base_date, t)
        except ValueError:
            continue
    raise ValueError(f"Invalid time format: {value}")


def _format_logged_time(total_minutes: int) -> str:
    hrs = total_minutes // 60
    mins = total_minutes % 60
    if hrs > 0 and mins > 0:
        return f"{hrs} hrs {mins} mins"
    if hrs > 0:
        return f"{hrs} hrs"
    return f"{mins} mins"


def _get_employee_in_company(employee_id: int) -> Employee:
    emp = Employee.query.get(employee_id)
    if not emp or emp.company_id != g.user.company_id:
        return None
    return emp


def _upsert_attendance(company_id: int, employee_id: int, att_date: date, payload: dict, capture_method: str):
    row = Attendance.query.filter_by(
        company_id=company_id,
        employee_id=employee_id,
        attendance_date=att_date
    ).first()

    is_new = False
    if not row:
        row = Attendance(company_id=company_id, employee_id=employee_id, attendance_date=att_date)
        row.created_by = getattr(g.user, "id", None)
        db.session.add(row)
        is_new = True

    # Update allowed fields
    row.status = payload.get("status", row.status) or row.status
    row.capture_method = capture_method

    login_at = payload.get("login_at")
    logout_at = payload.get("logout_at")

    if login_at is not None:
        row.punch_in_time = login_at
    if logout_at is not None:
        row.punch_out_time = logout_at

    row.updated_by = getattr(g.user, "id", None)

    # Auto calculate logged time
    row.recalc_total_minutes()

    return row, is_new


# -----------------------------
# 1) List Attendance (Table)
# -----------------------------
@attendance_bp.route("", methods=["GET"])
@token_required
@role_required(ALLOWED_MANAGE_ROLES)
def list_attendance():
    """
    Filters used by your UI:
      - role (Admin/HR/Manager/Employee/Accountant)
      - department
      - day=all/today
      - month (1-12) optional
      - from_date, to_date
      - search (name/email)
    """
    company_id = g.user.company_id

    role = request.args.get("role")          # Role dropdown
    department = request.args.get("department")
    day = (request.args.get("day") or "all").lower()  # all / today
    month = request.args.get("month")        # 1..12
    from_date = request.args.get("from_date")
    to_date = request.args.get("to_date")
    search = (request.args.get("search") or "").strip().lower()

    q = Attendance.query.filter_by(company_id=company_id)

    # Day filter
    if day == "today":
        q = q.filter(Attendance.attendance_date == date.today())

    # Date range filter
    if from_date:
        q = q.filter(Attendance.attendance_date >= _parse_date(from_date))
    if to_date:
        q = q.filter(Attendance.attendance_date <= _parse_date(to_date))

    # Month filter (optional)
    if month:
        try:
            m = int(month)
            q = q.filter(db.extract("month", Attendance.attendance_date) == m)
        except:
            pass

    # Join with employee + user for role/department/search
    q = q.join(Employee, Employee.id == Attendance.employee_id)\
         .join(User, User.id == Employee.user_id)\
         .filter(Employee.company_id == company_id)

    if department:
        q = q.filter(Employee.department == department)

    if role:
        q = q.filter(User.role == role)

    if search:
        q = q.filter(
            db.or_(
                db.func.lower(Employee.first_name).like(f"%{search}%"),
                db.func.lower(Employee.last_name).like(f"%{search}%"),
                db.func.lower(User.email).like(f"%{search}%"),
                db.func.lower(getattr(Employee, "employee_id", "")).like(f"%{search}%")
            )
        )

    rows = q.order_by(Attendance.attendance_date.desc()).limit(500).all()

    output = []
    for r in rows:
        emp = Employee.query.get(r.employee_id)
        user = User.query.get(emp.user_id) if emp else None

        output.append({
            "attendance_id": r.attendance_id,
            "employee_id": r.employee_id,
            "name": f"{emp.first_name} {emp.last_name}" if emp else "",
            "role": user.role if user else None,
            "department": emp.department if emp else None,
            "status": r.status,
            "logged_time": _format_logged_time(r.total_minutes),
            "login_at": r.punch_in_time.strftime("%H:%M") if r.punch_in_time else None,
            "logout_at": r.punch_out_time.strftime("%H:%M") if r.punch_out_time else None,
            "date": r.attendance_date.strftime("%d/%m/%Y"),
        })

    return jsonify({"attendance": output}), 200


# -----------------------------
# 2) Manual Attendance (Create/Upsert)
# -----------------------------
@attendance_bp.route("/manual", methods=["POST"])
@token_required
@role_required(ALLOWED_MANAGE_ROLES)
def manual_attendance():
    """
    Manual button action (UPSERT):
      Required: employee_id, date
      Optional: status, login_at, logout_at
    """
    data = request.get_json() or {}

    employee_id = data.get("employee_id")
    att_date = data.get("date")  # "2025-10-02" or "02/10/2025"
    status = data.get("status", "Present")

    if not employee_id or not att_date:
        return jsonify({"message": "employee_id and date are required"}), 400

    emp = _get_employee_in_company(int(employee_id))
    if not emp:
        return jsonify({"message": "Employee not found"}), 404

    d = _parse_date(att_date)

    login_at = None
    logout_at = None

    # If absent -> times should be empty (allowed)
    if data.get("login_at"):
        login_at = _parse_time(data["login_at"], d)
    if data.get("logout_at"):
        logout_at = _parse_time(data["logout_at"], d)

    payload = {
        "status": status,
        "login_at": login_at,
        "logout_at": logout_at
    }

    row, is_new = _upsert_attendance(g.user.company_id, emp.id, d, payload, capture_method="Manual")
    db.session.commit()

    return jsonify({
        "message": "Attendance saved successfully",
        "action": "inserted" if is_new else "updated",
        "attendance_id": row.attendance_id
    }), 200


# -----------------------------
# 3) Update Attendance (Edit icon)
# -----------------------------
@attendance_bp.route("/<int:attendance_id>", methods=["PUT"])
@token_required
@role_required(ALLOWED_MANAGE_ROLES)
def update_attendance(attendance_id):
    data = request.get_json() or {}
    row = Attendance.query.get(attendance_id)

    if not row or row.company_id != g.user.company_id:
        return jsonify({"message": "Attendance not found"}), 404

    # Update values
    if "status" in data:
        row.status = data["status"] or row.status

    if "login_at" in data:
        if data["login_at"]:
            row.punch_in_time = _parse_time(data["login_at"], row.attendance_date)
        else:
            row.punch_in_time = None

    if "logout_at" in data:
        if data["logout_at"]:
            row.punch_out_time = _parse_time(data["logout_at"], row.attendance_date)
        else:
            row.punch_out_time = None

    row.capture_method = "Manual"
    row.updated_by = getattr(g.user, "id", None)
    row.recalc_total_minutes()

    db.session.commit()
    return jsonify({"message": "Attendance updated successfully"}), 200


# -----------------------------
# 4) Delete Attendance (Delete icon)
# -----------------------------
@attendance_bp.route("/<int:attendance_id>", methods=["DELETE"])
@token_required
@role_required(ALLOWED_MANAGE_ROLES)
def delete_attendance(attendance_id):
    row = Attendance.query.get(attendance_id)
    if not row or row.company_id != g.user.company_id:
        return jsonify({"message": "Attendance not found"}), 404

    db.session.delete(row)
    db.session.commit()
    return jsonify({"message": "Attendance deleted successfully"}), 200


# -----------------------------
# 5) Import Attendance (CSV/XLSX) with UPSERT
# -----------------------------
@attendance_bp.route("/import", methods=["POST"])
@token_required
@role_required(ALLOWED_MANAGE_ROLES)
def import_attendance():
    """
    Supports:
      - CSV
      - XLSX (optional)
    Columns accepted:
      employee_id (required) OR employee_code (if you store it in Employee.employee_id string)
      date (required)
      status (Present/Absent)
      login_at (HH:MM)
      logout_at (HH:MM)

    UPSERT:
      company_id + employee_id + date exists => update
    """
    if "file" not in request.files:
        return jsonify({"message": "file is required (multipart/form-data)"}), 400

    f = request.files["file"]
    filename = (f.filename or "").lower()

    inserted = 0
    updated = 0
    errors = []

    company_id = g.user.company_id

    def handle_row(raw, row_no):
        nonlocal inserted, updated, errors

        try:
            # employee_id required
            emp_id = raw.get("employee_id")
            if not emp_id:
                raise ValueError("employee_id is required")

            emp = _get_employee_in_company(int(emp_id))
            if not emp:
                raise ValueError(f"Employee not found or not in company (employee_id={emp_id})")

            d = _parse_date(raw.get("date", ""))

            status = raw.get("status", "Present") or "Present"

            login_at = _parse_time(raw.get("login_at", ""), d) if raw.get("login_at") else None
            logout_at = _parse_time(raw.get("logout_at", ""), d) if raw.get("logout_at") else None

            payload = {"status": status, "login_at": login_at, "logout_at": logout_at}

            row, is_new = _upsert_attendance(company_id, emp.id, d, payload, capture_method="Import")
            if is_new:
                inserted += 1
            else:
                updated += 1

        except Exception as e:
            errors.append({"row": row_no, "error": str(e), "data": raw})

    # CSV
    if filename.endswith(".csv"):
        stream = io.StringIO(f.stream.read().decode("utf-8", errors="ignore"))
        reader = csv.DictReader(stream)

        for idx, r in enumerate(reader, start=2):  # 2 => header is row 1
            # normalize keys
            row = { (k or "").strip().lower(): (v or "").strip() for k, v in r.items() }
            handle_row(row, idx)

        db.session.commit()
        return jsonify({
            "message": "Import completed",
            "inserted": inserted,
            "updated": updated,
            "errors_count": len(errors),
            "errors": errors[:50]  # return first 50 only
        }), 200

    # XLSX (optional)
    if filename.endswith(".xlsx"):
        try:
            import openpyxl
        except ImportError:
            return jsonify({"message": "openpyxl not installed. Install it or use CSV."}), 400

        wb = openpyxl.load_workbook(f, data_only=True)
        ws = wb.active

        headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]

        for i in range(2, ws.max_row + 1):
            row_obj = {}
            for j, h in enumerate(headers, start=1):
                val = ws.cell(row=i, column=j).value
                row_obj[h] = str(val).strip() if val is not None else ""

            handle_row(row_obj, i)

        db.session.commit()
        return jsonify({
            "message": "Import completed",
            "inserted": inserted,
            "updated": updated,
            "errors_count": len(errors),
            "errors": errors[:50]
        }), 200

    return jsonify({"message": "Unsupported file type. Use .csv or .xlsx"}), 400


# -----------------------------
# 6) Employee view only (No create)
# -----------------------------
@attendance_bp.route("/me", methods=["GET"])
@token_required
@role_required(["EMPLOYEE"])
def my_attendance():
    """
    Employee can only view their own attendance.
    No punch in/out APIs here.
    """
    emp = Employee.query.filter_by(user_id=g.user.id, company_id=g.user.company_id).first()
    if not emp:
        return jsonify({"message": "Employee profile not found"}), 404

    q = Attendance.query.filter_by(company_id=g.user.company_id, employee_id=emp.id)\\
                        .order_by(Attendance.attendance_date.desc())\\
                        .limit(180)

    output = []
    for r in q.all():
        output.append({
            "status": r.status,
            "logged_time": _format_logged_time(r.total_minutes),
            "login_at": r.punch_in_time.strftime("%H:%M") if r.punch_in_time else None,
            "logout_at": r.punch_out_time.strftime("%H:%M") if r.punch_out_time else None,
            "date": r.attendance_date.strftime("%d/%m/%Y"),
        })

    return jsonify({"attendance": output}), 200
'''

routes_employee_advanced_py_content = "from flask import Blueprint\nemployee_advanced_bp = Blueprint('employee_advanced', __name__)"
routes_urls_py_content = "from flask import Blueprint\nurls_bp = Blueprint('urls', __name__)"
routes_permissions_py_content = "from flask import Blueprint\npermissions_bp = Blueprint('permissions', __name__)"

routes_documents_py_content = """
import os
from flask import Blueprint, request, jsonify, send_file, current_app, g
from werkzeug.utils import secure_filename
from models import db
from models.employee_documents import EmployeeDocument, DocumentType
from models.employee import Employee
from utils.decorators import token_required
import uuid
from datetime import datetime

documents_bp = Blueprint('documents', __name__)

def get_upload_folder():
    upload_folder = os.path.join(current_app.root_path, 'uploads', 'documents')
    os.makedirs(upload_folder, exist_ok=True)
    return upload_folder

@documents_bp.route('/upload', methods=['POST'])
@token_required
def upload_document():
    if 'document_file' not in request.files:
        return jsonify({'message': 'No file part'}), 400
    file = request.files['document_file']
    if file.filename == '':
        return jsonify({'message': 'No selected file'}), 400

    original_filename = secure_filename(file.filename)
    file_extension = original_filename.rsplit('.', 1)[1].lower()
    unique_filename = f"{uuid.uuid4().hex}.{file_extension}"
    upload_folder = get_upload_folder()
    file_path = os.path.join(upload_folder, unique_filename)
    file.save(file_path)

    new_document = EmployeeDocument(
        employee_id=g.user.employee_profile.id,
        document_type=request.form.get('document_type'),
        document_name=original_filename,
        file_path=file_path,
        file_url=f"/api/documents/download/{unique_filename}"
    )
    db.session.add(new_document)
    db.session.commit()
    return jsonify({'message': 'Document uploaded successfully'}), 201

@documents_bp.route('/download/<filename>')
@token_required
def download_document(filename):
    document = EmployeeDocument.query.filter(EmployeeDocument.file_url.endswith(filename)).first_or_404()
    if document.employee_id != g.user.employee_profile.id and g.user.role not in ['ADMIN', 'HR', 'SUPER_ADMIN']:
        return jsonify({'message': 'Access denied'}), 403
    return send_file(document.file_path, as_attachment=True)
"""

leave_init_py_content = "from flask import Blueprint\nleave_bp = Blueprint('leave', __name__, url_prefix='/leave')\nfrom . import routes"

leave_models_py_content = """
from datetime import datetime
from models import db

class LeaveType(db.Model):
    __tablename__ = 'leave_types'
    id = db.Column(db.Integer, primary_key=True)
    company_id = db.Column(db.Integer, db.ForeignKey('companies.id'))
    name = db.Column(db.String(100), nullable=False) # CL/SL/PL/WFH
    rules_json = db.Column(db.Text) # JSON string for rules
    leaves = db.relationship('LeaveRequest', backref='leave_type', lazy=True)

class LeaveRequest(db.Model):
    __tablename__ = 'leave_requests'
    id = db.Column(db.Integer, primary_key=True)
    employee_id = db.Column(db.Integer, db.ForeignKey('employees.id'), nullable=False)
    company_id = db.Column(db.Integer, db.ForeignKey('companies.id'), nullable=False)
    leave_type_id = db.Column(db.Integer, db.ForeignKey('leave_types.id'), nullable=False)
    from_date = db.Column(db.Date, nullable=False)
    to_date = db.Column(db.Date, nullable=False)
    reason = db.Column(db.Text)
    status = db.Column(db.String(50), default='Pending') # Pending/Approved/Rejected
    approved_by = db.Column(db.Integer, db.ForeignKey('employees.id'))
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    employee = db.relationship('Employee', foreign_keys=[employee_id])

class LeaveBalance(db.Model):
    __tablename__ = 'leave_balances'
    id = db.Column(db.Integer, primary_key=True)
    employee_id = db.Column(db.Integer, db.ForeignKey('employees.id'), nullable=False)
    leave_type_id = db.Column(db.Integer, db.ForeignKey('leave_types.id'), nullable=False)
    balance = db.Column(db.Float, default=0.0)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)
    
    employee = db.relationship('Employee', backref='leave_balances')
    leave_type = db.relationship('LeaveType', backref='balances')
"""

leave_routes_py_content = """
from flask import request, jsonify, g
from utils.decorators import token_required, role_required
from . import leave_bp
from .models import LeaveRequest
from models import db
from models.employee import Employee
from datetime import datetime

@leave_bp.route('/apply', methods=['POST'])
@token_required
def apply_leave():
    data = request.get_json()
    emp = Employee.query.filter_by(user_id=g.user.id).first()
    if not emp:
        return jsonify({'message': 'Employee profile not found'}), 404

    new_leave = LeaveRequest(
        employee_id=emp.id,
        company_id=emp.company_id,
        leave_type_id=data['leave_type_id'],
        from_date=datetime.strptime(data['from_date'], '%Y-%m-%d').date(),
        to_date=datetime.strptime(data['to_date'], '%Y-%m-%d').date(),
        reason=data['reason']
    )
    db.session.add(new_leave)
    db.session.commit()
    return jsonify({'message': 'Leave application submitted'}), 201

@leave_bp.route('/<int:id>/approve', methods=['POST'])
@token_required
@role_required(['HR', 'MANAGER', 'ADMIN'])
def approve_leave(id):
    leave = LeaveRequest.query.get_or_404(id)
    
    approver_emp = Employee.query.filter_by(user_id=g.user.id).first()
    
    data = request.get_json()
    leave.status = data.get('status', 'Approved')
    leave.approved_by = approver_emp.id if approver_emp else None
    db.session.commit()
    return jsonify({'message': f'Leave {leave.status.lower()}'})
"""

utils_decorators_py_content = """
from flask import request, jsonify, g
import jwt
from functools import wraps
from config import Config
from models.user import User

def token_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        token = None
        if 'Authorization' in request.headers:
            auth_header = request.headers['Authorization']
            if auth_header.startswith('Bearer '):
                token = auth_header.split(' ')[1]
        if not token:
            return jsonify({'message': 'Token is missing'}), 401
        try:
            data = jwt.decode(token, Config.SECRET_KEY, algorithms=["HS256"])
            g.user = User.query.get(data['user_id'])
            if not g.user:
                return jsonify({'message': 'User not found'}), 401
        except Exception as e:
            return jsonify({'message': 'Token is invalid!', 'error': str(e)}), 401
        return f(*args, **kwargs)
    return decorated

def role_required(allowed_roles):
    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            if g.user.role not in allowed_roles:
                return jsonify({'message': 'Permission denied'}), 403
            return f(*args, **kwargs)
        return decorated_function
    return decorator

def permission_required(permission_code):
    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            if not g.user.has_permission(permission_code):
                return jsonify({'message': 'Insufficient permissions'}), 403
            return f(*args, **kwargs)
        return decorated_function
    return decorator
"""

utils_url_generator_py_content = """
from flask import current_app

def clean_domain(s: str) -> str:
    if not s:
        return ""
    return s.replace("http://", "").replace("https://", "").strip().strip("/")

ROOT_DOMAIN = "company.com"

def build_web_address(subdomain: str) -> str:
    sub = clean_domain(subdomain)
    if not sub:
        return "localhost:5173"
    return f"{sub}.{ROOT_DOMAIN}"

def build_common_login_url(subdomain: str) -> str:
    sub = clean_domain(subdomain)
    if not sub:
        return "http://localhost:5173/login"
    return f"https://{sub}.{ROOT_DOMAIN}/login"

def build_company_base_url(subdomain: str) -> str:
    sub = (subdomain or "").strip().lower()
    if not sub:
        return current_app.config.get("FRONTEND_LOCAL", "http://localhost:5173")

    protocol = current_app.config.get("FRONTEND_PROTOCOL", "https")
    base_domain = current_app.config.get("FRONTEND_BASE_DOMAIN", "company.com")
    return f"{protocol}://{sub}.{base_domain}"
"""

utils_email_utils_py_content = """
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from flask import current_app
from datetime import datetime

def _send_plain_email(to_email: str, subject: str, body: str) -> bool:
    smtp_server = current_app.config.get("MAIL_SERVER", "smtp.gmail.com")
    smtp_port = int(current_app.config.get("MAIL_PORT", 587))
    smtp_user = current_app.config.get("MAIL_USERNAME")
    smtp_pass = current_app.config.get("MAIL_PASSWORD")
    sender = current_app.config.get("MAIL_DEFAULT_SENDER", smtp_user)

    if not smtp_user or not smtp_pass:
        print(f"❌ Mail credentials missing. Mock sending to {to_email}")
        return False

    try:
        msg = MIMEMultipart()
        msg["From"] = f"HRMS Team <{sender}>"
        msg["To"] = to_email
        msg["Subject"] = subject
        msg.attach(MIMEText(body, "plain"))

        server = smtplib.SMTP(smtp_server, smtp_port, timeout=50)
        server.ehlo()
        server.starttls()
        server.login(smtp_user, smtp_pass)
        server.send_message(msg)
        server.quit()

        print(f"✅ Email sent to {to_email}")
        return True
    except Exception as e:
        print(f"❌ Email error: {e}")
        return False

def send_signup_otp(to_email: str, otp: str) -> bool:
    subject = "Super Admin Signup OTP"
    body = f"Your signup OTP is: {otp}\\n\\nValid for 10 minutes."
    return _send_plain_email(to_email, subject, body)

def send_account_created_alert(personal_email: str, company_name: str, created_by: str) -> bool:
    subject = "Account Created Alert"
    body = (
        "Hello,\\n\\n"
        f"An account was created for you in {company_name} by {created_by}.\\n"
        f"Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\\n\\n"
        f"Regards,\\n"
        f"{company_name}\\n"
    )
    return _send_plain_email(personal_email, subject, body)

def send_login_credentials(personal_email: str, company_email: str, password: str,
                           company_name: str, web_address: str, login_url: str, created_by: str) -> bool:
    subject = "Login Details"
    body = (
        "Hello,\\n\\n"
        f"Your account has been created by {created_by}.\\n\\n"
        "Login Details:\\n"
        f"Web Address: {web_address}\\n"
        f"Username: {company_email}\\n"
        f"Password: {password}\\n\\n"
        "👉 Click here to login:\\n"
        f"{login_url}\\n\\n"
        f"Regards,\\n"
        f"{company_name}\\n"
    )
    return _send_plain_email(personal_email, subject, body)

def send_login_success_email(to_email: str) -> bool:
    subject = "Login Successful"
    body = (
        "Hello,\\n\\n"
        "Login successful.\\n"
        f"Date & Time: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\\n\\n"
        f"Regards,\\n"
        "Company Team\\n"
    )
    return _send_plain_email(to_email, subject, body)
"""

rebuild_db_py_content = """
from app import app
from models import db

if __name__ == "__main__":
    with app.app_context():
        print("🗑️  Dropping all tables...")
        db.drop_all()
        print("🔨 Creating all tables...")
        db.create_all()
        print("✅ Database rebuilt successfully!")
"""

clear_super_admin_py_content = """
from app import app
from models import db
from models.user import User
from models.employee import Employee

if __name__ == "__main__":
    with app.app_context():
        print("🧹 Clearing Super Admin data...")
        try:
            super_admins = User.query.filter_by(role='SUPER_ADMIN').all()
            for user in super_admins:
                Employee.query.filter_by(user_id=user.id).delete()
                db.session.delete(user)
            db.session.commit()
            print(f"✅ Cleared {len(super_admins)} Super Admin(s).")
        except Exception as e:
            db.session.rollback()
            print(f"❌ Error: {e}")
"""

env_content = """
SECRET_KEY=hrms-secret-key-change-this
MAIL_USERNAME=your-email@gmail.com
MAIL_PASSWORD=your-app-password
MAIL_DEFAULT_SENDER=your-email@gmail.com
"""

# ==============================================================================
# MAIN ORGANIZATION LOGIC
# ==============================================================================

def organize_project():
    """
    Cleans up the project directory by creating a standard structure,
    writing the correct code to files, and deleting obsolete/misplaced files.
    """
    print("🚀 Starting project organization...")

    # 1. Define the desired file structure and content
    project_files = {
        "app.py": app_py_content,
        "config.py": config_py_content,
        "routes/__init__.py": "",
        "routes/auth.py": routes_auth_py_content,
        "routes/admin.py": routes_admin_py_content,
        "routes/superadmin.py": routes_superadmin_py_content,
        "routes/hr.py": routes_hr_py_content,
        "routes/employee.py": routes_employee_py_content,
        "routes/attendance.py": routes_attendance_py_content,
        "routes/employee_advanced.py": routes_employee_advanced_py_content,
        "routes/urls.py": routes_urls_py_content,
        "routes/permissions.py": routes_permissions_py_content,
        "routes/documents.py": routes_documents_py_content,
        "models/__init__.py": models_init_py_content,
        "models/user.py": models_user_py_content,
        "models/company.py": models_company_py_content,
        "models/employee.py": models_employee_py_content,
        "models/attendance.py": models_attendance_py_content,
        "models/permission.py": models_permission_py_content,
        "models/department.py": models_department_py_content,
        "models/filter.py": models_filter_py_content,
        "models/urls.py": models_urls_py_content,
        "models/payroll.py": models_payroll_py_content,
        "models/employee_address.py": models_employee_address_py_content,
        "models/employee_bank.py": models_employee_bank_py_content,
        "models/employee_documents.py": models_employee_documents_py_content,
        "leave/__init__.py": leave_init_py_content,
        "leave/models.py": leave_models_py_content,
        "leave/routes.py": leave_routes_py_content,
        "utils/__init__.py": "",
        "utils/decorators.py": utils_decorators_py_content,
        "utils/url_generator.py": utils_url_generator_py_content,
        "utils/email_utils.py": utils_email_utils_py_content,
        "rebuild_db.py": rebuild_db_py_content,
        "clear_super_admin.py": clear_super_admin_py_content,
        ".env": env_content,
    }

    # 2. Create/update all the correct files
    print("\n--- Writing correct files ---")
    for path, content in project_files.items():
        if path == ".env" and os.path.exists(path):
            print(f"⚠️  Skipping .env (already exists)")
            continue
        create_file(path, content)

    # 3. Move all other root-level .py scripts to a 'scripts' directory
    print("\n--- Moving helper scripts ---")
    scripts_dir = "scripts"
    if not os.path.exists(scripts_dir):
        os.makedirs(scripts_dir)
    
    root_py_files = [f for f in os.listdir('.') if os.path.isfile(f) and f.endswith('.py')]
    core_app_files = ['app.py', 'config.py', 'organize_project.py', 'rebuild_db.py', 'clear_super_admin.py']

    for script in root_py_files:
        if script not in core_app_files:
            try:
                shutil.move(script, os.path.join(scripts_dir, script))
                print(f"🚚 Moved '{script}' to '{scripts_dir}/'")
            except Exception as e:
                print(f"❌ Could not move '{script}': {e}")

    # 4. Define and delete all obsolete/conflicting paths
    print("\n--- Deleting obsolete files and directories ---")
    obsolete_paths = [
        'auth', 
        'superadmin', 
        'employee',
        'admin',
        'routes.py', 
        'models.py', 
        '__init__.py',
        'models/admin.py',
        'models/hr.py',
        'models/routes.py',
        'models/permissions.py',
        'models/employee_advanced.py',
        'routes/documents_routes.py' # Old name
    ]
    for path in obsolete_paths:
        delete_path(path)

    print("\n✨ Project organization complete!")
    print("👉 You can now run 'python app.py' to start the server.")
    print("👉 It is safe to delete 'organize_project.py' now.")

if __name__ == '__main__':
    organize_project()